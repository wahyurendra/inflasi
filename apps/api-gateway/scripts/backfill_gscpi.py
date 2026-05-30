"""Backfill `ext_supply_chain_index` with NY Fed's GSCPI series.

GSCPI (Global Supply Chain Pressure Index) is published as a monthly XLSX on
the NY Fed research page. The dataset is small (~1 row/month since 1998), so
we just pull it whole on every run and upsert.

Source: https://www.newyorkfed.org/research/policy/gscpi

Two fetch modes:

* `--source remote` (default) — download the XLSX directly. Requires `openpyxl`.
* `--source csv --path /local/gscpi.csv` — load a local CSV when the operator
  has already downloaded one (e.g. behind a corporate proxy).

The CSV / XLSX shape NY Fed publishes:

    Date, GSCPI
    1998-01-01, 0.32
    1998-02-01, 0.30
    ...

This script accepts dates in YYYY-MM or YYYY-MM-DD form.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import io
import logging
import os
import sys
from datetime import date
from typing import Any, Iterable

import httpx
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

logger = logging.getLogger("backfill_gscpi")

# NY Fed has stable URLs but they version filenames occasionally — fail loud
# rather than silently fetching the wrong file.
DEFAULT_GSCPI_URL = "https://www.newyorkfed.org/medialibrary/research/interactives/gscpi/downloads/gscpi_data.xlsx"


async def load_remote_xlsx(url: str) -> list[tuple[date, float]]:
    """Download GSCPI workbook (XLS or XLSX) and parse rows. Pure I/O — no DB.

    NY Fed currently serves an XLS (legacy binary, OLE compound) file under
    `gscpi_data.xlsx` — the extension is misleading. We sniff the first few
    bytes and dispatch to xlrd for XLS, openpyxl for true XLSX (zip).
    """
    async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
        resp = await client.get(url)
        resp.raise_for_status()
        raw = resp.content

    head = raw[:4]
    if head[:2] == b"PK":
        return _parse_xlsx(raw)
    if head == b"\xd0\xcf\x11\xe0":
        return _parse_xls(raw)
    # Neither magic — try both, in best-effort order.
    try:
        return _parse_xlsx(raw)
    except Exception:
        return _parse_xls(raw)


def _parse_xlsx(raw: bytes) -> list[tuple[date, float]]:
    """Parse modern XLSX via openpyxl."""
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(
            "openpyxl is required for XLSX; pip install openpyxl"
        ) from exc

    wb = load_workbook(io.BytesIO(raw), data_only=True)
    rows: list[tuple[date, float]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, date_col_idx, value_col_idx = _find_header_xlsx(ws)
        if date_col_idx is None or value_col_idx is None:
            continue
        for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if date_col_idx >= len(r) or value_col_idx >= len(r):
                continue
            d = _coerce_date(r[date_col_idx])
            v = _coerce_float(r[value_col_idx])
            if d and v is not None:
                rows.append((d, v))
        if rows:
            break
    return rows


def _parse_xls(raw: bytes) -> list[tuple[date, float]]:
    """Parse legacy XLS via xlrd. NY Fed's published file is XLS as of 2026."""
    try:
        import xlrd
    except Exception as exc:
        raise RuntimeError(
            "xlrd is required for legacy XLS; pip install 'xlrd>=2.0'"
        ) from exc

    wb = xlrd.open_workbook(file_contents=raw)
    rows: list[tuple[date, float]] = []
    for sheet in wb.sheets():
        header_row, date_col_idx, value_col_idx = _find_header_xls(sheet)
        if date_col_idx is None or value_col_idx is None:
            continue
        for ri in range(header_row + 1, sheet.nrows):
            row = sheet.row_values(ri)
            if date_col_idx >= len(row) or value_col_idx >= len(row):
                continue
            d = _coerce_date(row[date_col_idx])
            v = _coerce_float(row[value_col_idx])
            if d and v is not None:
                rows.append((d, v))
        if rows:
            break
    return rows


def _find_header_xlsx(ws) -> tuple[int, int | None, int | None]:
    """Locate the (Date, GSCPI) header row inside an XLSX sheet."""
    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
        labels = [str(c).strip().lower() if c is not None else "" for c in row]
        if "date" in labels and any("gscpi" in lab for lab in labels):
            date_idx = labels.index("date")
            gscpi_idx = next(i for i, lab in enumerate(labels) if "gscpi" in lab)
            return ri, date_idx, gscpi_idx
        if ri > 20:
            break
    return 0, None, None


def _find_header_xls(sheet) -> tuple[int, int | None, int | None]:
    """Locate the (Date, GSCPI) header row inside an XLS sheet."""
    for ri in range(min(sheet.nrows, 20)):
        row = sheet.row_values(ri)
        labels = [str(c).strip().lower() if c is not None else "" for c in row]
        if "date" in labels and any("gscpi" in lab for lab in labels):
            date_idx = labels.index("date")
            gscpi_idx = next(i for i, lab in enumerate(labels) if "gscpi" in lab)
            return ri, date_idx, gscpi_idx
    return 0, None, None


def load_csv(path: str) -> list[tuple[date, float]]:
    rows: list[tuple[date, float]] = []
    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None) or []
        labels = [c.strip().lower() for c in header]
        try:
            date_idx = labels.index("date")
        except ValueError:
            date_idx = 0
        try:
            gscpi_idx = next(i for i, lab in enumerate(labels) if "gscpi" in lab)
        except StopIteration:
            gscpi_idx = 1
        for row in reader:
            if date_idx >= len(row) or gscpi_idx >= len(row):
                continue
            d = _coerce_date(row[date_idx])
            v = _coerce_float(row[gscpi_idx])
            if d and v is not None:
                rows.append((d, v))
    return rows


def _coerce_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, date):
        return v.replace(day=1)
    s = str(v).strip()
    if not s:
        return None
    # NY Fed's GSCPI XLS uses "31-Jan-1998" — include `%d-%b-%Y`.
    for fmt in ("%Y-%m-%d", "%Y-%m", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date().replace(day=1)
        except ValueError:
            continue
    return None


def _coerce_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


async def upsert(db: AsyncSession, rows: Iterable[tuple[date, float]]) -> int:
    written = 0
    for period, value in rows:
        await db.execute(
            text("""
                INSERT INTO ext_supply_chain_index (periode, gscpi)
                VALUES (:periode, :gscpi)
                ON CONFLICT (periode) DO UPDATE SET gscpi = EXCLUDED.gscpi
            """),
            {"periode": period, "gscpi": value},
        )
        written += 1
    await db.commit()
    return written


async def run(*, source: str, url: str, path: str | None) -> int:
    if source == "csv":
        if not path:
            raise SystemExit("--source csv requires --path")
        rows = load_csv(path)
    else:
        rows = await load_remote_xlsx(url)
    if not rows:
        logger.warning("no rows parsed from %s", path or url)
        return 0

    db_url = _resolve_url()
    engine = create_async_engine(db_url, pool_recycle=300)
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    try:
        async with factory() as db:
            written = await upsert(db, rows)
    finally:
        await engine.dispose()
    logger.info("wrote %s GSCPI rows (range %s..%s)",
                written, rows[0][0], rows[-1][0])
    return written


def _resolve_url() -> str:
    raw = os.getenv("ANALYTICS_DATABASE_URL") or os.getenv("DATABASE_URL") or ""
    if raw.startswith("postgresql://"):
        return raw.replace("postgresql://", "postgresql+asyncpg://", 1)
    return raw or "postgresql+asyncpg://postgres:postgres@localhost:5432/inflasi"


def _cli() -> None:
    parser = argparse.ArgumentParser(description="Backfill NY Fed GSCPI into ext_supply_chain_index.")
    parser.add_argument("--source", choices=["remote", "csv"], default="remote")
    parser.add_argument("--url", default=DEFAULT_GSCPI_URL)
    parser.add_argument("--path", default=None, help="Local CSV path (with --source csv)")
    parser.add_argument("--log-level", default="INFO")
    args = parser.parse_args()
    logging.basicConfig(
        level=args.log_level.upper(),
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )
    asyncio.run(run(source=args.source, url=args.url, path=args.path))


if __name__ == "__main__":
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    _cli()
