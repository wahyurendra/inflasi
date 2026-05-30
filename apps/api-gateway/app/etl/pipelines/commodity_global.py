"""Pipeline: Global Commodity Prices (World Bank Pink Sheet — Monthly XLSX)

Source: https://www.worldbank.org/en/research/commodity-markets
File:   CMO-Historical-Data-Monthly.xlsx (published roughly monthly)

The old World Bank GEM Commodities API path (`/v2/country/WLD/indicator/RICE_05`)
returns "Invalid value" because the GEM source no longer exposes commodity
series under that endpoint. The XLSX bulk file is now the canonical interface.

Layout of the "Monthly Prices" sheet:

    r1-r4: title / copyright / "updated on ..."
    r5:    full name ("Crude oil, average ...")
    r6:    unit ("($/bbl)")
    r7:    indicator code ("CRUDE_BRENT", "RICE_05", ...)
    r8+:   data — column A = "YYYYMmm" (e.g. "2024M05"), columns B+ = values

We grab the indicator codes we care about, then walk data rows and emit one
`ext_commodity_price` row per (period, commodity) pair.
"""

from __future__ import annotations

import io
import logging
from datetime import date

import httpx
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger(__name__)

PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)

# (WB indicator code in row 7 → (internal commodity name, unit)).
# Stay close to the original pipeline's mapping so downstream queries don't
# need to learn new names.
WB_INDICATORS: dict[str, tuple[str, str]] = {
    "RICE_05": ("rice", "USD/mt"),
    "WHEAT_US_HRW": ("wheat", "USD/mt"),
    "MAIZE": ("corn", "USD/mt"),
    "PALM_OIL": ("palm_oil", "USD/mt"),
    "SOYBEAN_OIL": ("soybean_oil", "USD/mt"),
    "SUGAR_WLD": ("sugar", "USD/kg"),
    "UREA_EE_BULK": ("urea", "USD/mt"),
    "CRUDE_BRENT": ("crude_oil_brent", "USD/bbl"),
}


class CommodityGlobalPipeline:
    """Ingest the World Bank Pink Sheet into `ext_commodity_price`."""

    name = "commodity_global"

    def __init__(
        self,
        db: AsyncSession,
        *,
        year_start: int = 2020,
        url: str = PINK_SHEET_URL,
    ):
        self.db = db
        self.year_start = year_start
        self.year_end = date.today().year
        self.url = url
        self.client = httpx.AsyncClient(timeout=60.0, follow_redirects=True)

    async def run(self) -> int:
        logger.info(
            "[%s] downloading Pink Sheet (years %s..%s)",
            self.name, self.year_start, self.year_end,
        )
        try:
            content = await self._download()
            if not content:
                return 0
            rows = self._parse(content)
            if not rows:
                logger.warning("[%s] no rows parsed", self.name)
                return 0
            count = await self._load(rows)
            logger.info("[%s] loaded %s records", self.name, count)
            return count
        finally:
            await self.client.aclose()

    # ── Download ──────────────────────────────────────────────

    async def _download(self) -> bytes | None:
        r = await self.client.get(self.url)
        if r.status_code != 200:
            logger.warning("[%s] HTTP %s on %s", self.name, r.status_code, self.url)
            return None
        return r.content

    # ── Parse ─────────────────────────────────────────────────

    def _parse(self, content: bytes) -> list[dict]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            logger.error("[%s] openpyxl unavailable: %s", self.name, exc)
            return []

        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        if "Monthly Prices" not in wb.sheetnames:
            logger.warning("[%s] sheet 'Monthly Prices' missing", self.name)
            return []
        ws = wb["Monthly Prices"]

        # Stream rows; track the code row + data once we hit it.
        code_to_col: dict[str, int] = {}
        out: list[dict] = []
        seen_code_row = False

        for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not seen_code_row:
                if ri < 5 or row is None:
                    continue
                # Indicator codes live somewhere around row 7; the first row
                # whose cells include any known WB code is the code row.
                if any(
                    isinstance(c, str) and c in WB_INDICATORS
                    for c in row if c is not None
                ):
                    for ci, cell in enumerate(row):
                        if isinstance(cell, str) and cell in WB_INDICATORS:
                            code_to_col[cell] = ci
                    seen_code_row = True
                continue

            # Data row: column 0 is "YYYYMmm" or similar.
            if not row or row[0] is None:
                continue
            period = _parse_period(row[0])
            if period is None:
                continue
            if period.year < self.year_start or period.year > self.year_end:
                continue
            for code, (commodity, unit) in WB_INDICATORS.items():
                col = code_to_col.get(code)
                if col is None or col >= len(row):
                    continue
                value = row[col]
                if value is None:
                    continue
                try:
                    price = float(value)
                except (TypeError, ValueError):
                    continue
                if price <= 0:
                    continue
                out.append({
                    "periode": period,
                    "commodity": commodity,
                    "price": price,
                    "unit": unit,
                })

        wb.close()
        return out

    # ── Load ──────────────────────────────────────────────────

    async def _load(self, rows: list[dict]) -> int:
        rows.sort(key=lambda x: (x["commodity"], x["periode"]))
        prev: dict[str, float] = {}
        count = 0
        for row in rows:
            commodity = row["commodity"]
            price = row["price"]
            last = prev.get(commodity)
            change_pct = (
                round((price - last) / last * 100, 4)
                if last and last > 0 else None
            )
            prev[commodity] = price
            await self.db.execute(
                text("""
                    INSERT INTO ext_commodity_price
                      (periode, commodity, price, unit, change_pct, sumber)
                    VALUES (:periode, :commodity, :price, :unit, :change_pct, 'WORLD_BANK')
                    ON CONFLICT (periode, commodity)
                    DO UPDATE SET
                      price = EXCLUDED.price,
                      change_pct = EXCLUDED.change_pct
                """),
                {**row, "change_pct": change_pct},
            )
            count += 1
        await self.db.commit()
        return count


# ── Helpers ──────────────────────────────────────────────────

def _parse_period(value) -> date | None:
    """Pink Sheet period column is `YYYYMmm` (e.g. "2024M05"). Also tolerate
    datetimes / plain `YYYY-MM` for safety against future format tweaks."""
    if value is None:
        return None
    if hasattr(value, "year") and hasattr(value, "month"):
        return date(value.year, value.month, 1)
    s = str(value).strip()
    if "M" in s and len(s) >= 6:
        try:
            year_part, month_part = s.split("M", 1)
            year = int(year_part)
            month = int(month_part)
            if 1900 < year < 2100 and 1 <= month <= 12:
                return date(year, month, 1)
        except ValueError:
            pass
    if len(s) >= 7 and "-" in s:
        try:
            return date.fromisoformat(s[:7] + "-01")
        except ValueError:
            pass
    return None
